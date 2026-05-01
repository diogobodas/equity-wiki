"""
Build interactive xlsx with SmartFit KPIs + 1T26 estimate.

Modelo: receita Brasil é OUTPUT das premissas operacionais.
Inputs: # academias EoP, # alunos EoP, ticket ex-TP, TP freq%, TP rev%.
Receita_total_BR = (alunos_média × ticket_ex_TP × 3) / (1 - tp_rev_pct) / 1000.
"""
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
estimate_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
green_out = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
bold = Font(bold=True)
italic_gray = Font(italic=True, color="595959")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# === Sheet 1: Premissas ===
ws = wb.active
ws.title = "Premissas"

ws["A1"] = "PREMISSAS — 1T26 SmartFit Brasil"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:C1")

ws["A3"] = "INPUTS OPERACIONAIS (amarelo = editável)"
ws["A3"].font = bold
ws["A3"].fill = gray
ws.merge_cells("A3:C3")

# Receita NÃO é input — é output. Inputs = drivers operacionais.
inputs_op = [
    ("# academias Smart Fit Brasil próprias EoP", 720, "+27 net adds vs 4T25 (1T tipicamente menor que 4T)"),
    ("# alunos Smart Fit Brasil próprias EoP (mil)", 1_818, "+6% YoY (desaceleração; 1T25 foi +12% YoY vs 1T24)"),
    ("Ticket ex-TP 1T26 (R$/mês — balcão puro)", 105.0, "1T25 ex-TP foi 104.6; +0.4% YoY (repasse Black novos cancela vs 1T25)"),
]
for i, (label, val, note) in enumerate(inputs_op, start=4):
    ws[f"A{i}"] = label
    ws[f"B{i}"] = val
    ws[f"B{i}"].fill = yellow
    ws[f"B{i}"].border = border
    if "Ticket" in label:
        ws[f"B{i}"].number_format = "#,##0.0"
    else:
        ws[f"B{i}"].number_format = "#,##0"
    ws[f"C{i}"] = note
    ws[f"C{i}"].font = italic_gray

ws["A8"] = ""
ws["A9"] = "INPUTS TOTALPASS (editáveis)"
ws["A9"].font = bold
ws["A9"].fill = gray
ws.merge_cells("A9:C9")

inputs_tp = [
    ("TP freq% 1T26 (% check-ins TP em SF próprias BR)", 0.165, "Default: 2026 anual 18% × peso MAU 1T26 × intensidade 1.10 = 16.5%"),
    ("TP rev% 1T26 (% receita SF própria BR via TP)", 0.138, "Default: 2026 anual 15% × mesmo peso = 13.8%"),
]
for i, (label, val, note) in enumerate(inputs_tp, start=10):
    ws[f"A{i}"] = label
    ws[f"B{i}"] = val
    ws[f"B{i}"].number_format = "0.0%"
    ws[f"B{i}"].fill = yellow
    ws[f"B{i}"].border = border
    ws[f"C{i}"] = note
    ws[f"C{i}"].font = italic_gray

# Outputs DERIVED (read-only)
ws["A13"] = ""
ws["A14"] = "OUTPUTS DERIVADOS (verde = formula → muda quando você ajusta os inputs)"
ws["A14"].font = bold
ws["A14"].fill = gray
ws.merge_cells("A14:C14")

# Need to get prior period: 4T25 EoP acad = 693, alunos = 1595
ws["A15"] = "Alunos média 1T26 (mil)"
ws["B15"] = "=(B5+1595)/2"
ws["B15"].number_format = "#,##0"
ws["B15"].fill = green_out
ws["B15"].border = border
ws["C15"] = "= (alunos_EoP_1T26 + alunos_EoP_4T25=1595) / 2"
ws["C15"].font = italic_gray

ws["A16"] = "# acad média 1T26"
ws["B16"] = "=(B4+693)/2"
ws["B16"].number_format = "#,##0.0"
ws["B16"].fill = green_out
ws["B16"].border = border
ws["C16"] = "= (acad_EoP_1T26 + acad_EoP_4T25=693) / 2"
ws["C16"].font = italic_gray

ws["A17"] = "Receita balcão 1T26 (R$ mm)"
ws["B17"] = "=B6*B15*3/1000"
ws["B17"].number_format = "#,##0.0"
ws["B17"].fill = green_out
ws["B17"].border = border
ws["C17"] = "= ticket_ex_TP × alunos_média × 3 meses / 1000"
ws["C17"].font = italic_gray

ws["A18"] = "Receita TOTAL Smart Fit Brasil 1T26 (R$ mm)"
ws["B18"] = "=B17/(1-B11)"
ws["B18"].number_format = "#,##0.0"
ws["B18"].fill = green_out
ws["B18"].border = border
ws["B18"].font = bold
ws["C18"] = "= receita_balcão / (1 - TP_rev%) — ESTE é o número comparável ao disclosed"
ws["C18"].font = italic_gray

ws["A19"] = "Ticket consolidado 1T26 (R$/mês — disclosed)"
ws["B19"] = "=B6/(1-B11)"
ws["B19"].number_format = "#,##0.0"
ws["B19"].fill = green_out
ws["B19"].border = border
ws["C19"] = "= ticket_ex_TP / (1 - TP_rev%) — comparável ao ticket reportado"
ws["C19"].font = italic_gray

# YoY referência
ws["A20"] = ""
ws["A21"] = "REFERÊNCIA — 1T25 (anchor YoY)"
ws["A21"].font = bold
ws["A21"].fill = gray
ws.merge_cells("A21:C21")

ref_1t25 = [
    ("Receita Smart Fit Brasil 1T25", 577.5, ""),
    ("# academias EoP 1T25", 573, ""),
    ("# alunos EoP 1T25", 1_715, ""),
    ("Ticket ex-TP 1T25", 104.6, ""),
    ("Ticket consolidado 1T25", 117.6, ""),
    ("TP freq% 1T25", 0.138, ""),
    ("TP rev% 1T25", 0.110, ""),
]
for i, (label, val, _) in enumerate(ref_1t25, start=22):
    ws[f"A{i}"] = label
    ws[f"B{i}"] = val
    if isinstance(val, float) and val < 1:
        ws[f"B{i}"].number_format = "0.0%"
    elif "Ticket" in label or "Receita" in label:
        ws[f"B{i}"].number_format = "#,##0.0"
    else:
        ws[f"B{i}"].number_format = "#,##0"

ws["A29"] = ""
ws["A30"] = "TP MAU Brasil — Sensor Tower (referência)"
ws["A30"].font = bold
ws["A30"].fill = gray
ws.merge_cells("A30:C30")

ref_mau = [
    ("TP MAU 1T26 (Jan-Mar)", 11_554_000, "Real"),
    ("TP MAU 1T25", 5_641_066, "+105% YoY MAU"),
    ("TP MAU 2025 anual", 27_312_195, "anchor 15% freq disclosed"),
    ("TP MAU 2024 anual", 14_646_891, "anchor 11% freq disclosed; growth 25/24 = +86%"),
]
for i, (label, val, note) in enumerate(ref_mau, start=31):
    ws[f"A{i}"] = label
    ws[f"B{i}"] = val
    ws[f"B{i}"].number_format = "#,##0"
    ws[f"C{i}"] = note
    ws[f"C{i}"].font = italic_gray

# Notas
ws["A36"] = "NOTAS"
ws["A36"].font = bold
ws["A36"].fill = gray
ws.merge_cells("A36:C36")

notas = [
    "• Receita NÃO é input direto — é OUTPUT de (alunos × ticket × 3) / (1 - TP_rev%).",
    "• Sensibilize ticket ex-TP YoY: 1T25 foi 104.6, default 1T26 = 105.0 (+0.4% YoY).",
    "  → Plano Black 1T26 tem repasse novo, mas o mesmo ocorreu em 1T25 (cancela YoY).",
    "  → 1T25 ticket ex-TP já foi +4.7% YoY vs 1T24 (efeito Black acumulado).",
    "• 4T25 ticket ex-TP foi 109.2 — levemente exagerado pelas 88 aberturas (alunos média baixa proporcional).",
    "• Sazonalidade uso 1T = 1.10× → TP freq% 1T26 acima do anual 18% × intensidade.",
    "• TP MAU 1T26 = 11.55M (+105% YoY). Se 2026 anual = 18% freq, 1T26 ≈ 16.5%.",
    "• Mude qualquer célula amarela e a Tabela KPIs recalcula tudo.",
]
for i, n in enumerate(notas, start=37):
    ws[f"A{i}"] = n
    ws[f"A{i}"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{i}:C{i}")

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 65

# === Sheet 2: Tabela KPIs ===
ws2 = wb.create_sheet("Tabela KPIs")

ws2["A1"] = "Smart Fit Brasil próprias — KPIs operacionais"
ws2["A1"].font = Font(bold=True, size=12)
ws2.merge_cells("A1:R1")
ws2["A2"] = "Histórico 1T23-4T25 + estimativa 1T26 (linha amarela = recalcula com Premissas)"
ws2["A2"].font = italic_gray
ws2.merge_cells("A2:R2")

headers = [
    "Período", "Receita BR (R$ mm)", "# acad EoP", "Alunos EoP (mil)",
    "# acad média", "Alunos média (mil)", "R/Loja anual (R$ mm)", "YoY",
    "Ticket mensal (R$)", "YoY", "Alunos/loja (mil)", "YoY",
    "Alunos/loja c/TP (mil)", "YoY", "Ticket ex-TP (R$)", "YoY",
    "TP freq%", "TP rev%",
]
for col_idx, h in enumerate(headers, start=1):
    cell = ws2.cell(row=4, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    cell.border = border

# TP shares trimestrais (do script com intensidade)
TP_BY_Q = {
    "1T24": (9.27, 6.74), "2T24": (9.78, 7.11),  "3T24": (12.74, 9.27), "4T24": (12.21, 8.88),
    "1T25": (13.76, 11.01), "2T25": (12.52, 10.02), "3T25": (16.31, 13.05), "4T25": (17.41, 13.93),
}

HIST_DATA = [
    ("4T22", None, 429, 1_165, None, None),
    ("1T23", 383.5, 431, 1_307, None, None),
    ("2T23", 405.7, 431, 1_277, None, None),
    ("3T23", 413.0, 448, 1_316, None, None),
    ("4T23", 425.1, 486, 1_353, None, None),
    ("1T24", 464.8, 493, 1_525, *TP_BY_Q["1T24"]),
    ("2T24", 482.0, 506, 1_515, *TP_BY_Q["2T24"]),
    ("3T24", 503.7, 525, 1_559, *TP_BY_Q["3T24"]),
    ("4T24", 524.0, 569, 1_560, *TP_BY_Q["4T24"]),
    ("1T25", 577.5, 573, 1_715, *TP_BY_Q["1T25"]),
    ("2T25", 595.7, 587, 1_635, *TP_BY_Q["2T25"]),
    ("3T25", 605.3, 605, 1_620, *TP_BY_Q["3T25"]),
    ("4T25", 611.7, 693, 1_595, *TP_BY_Q["4T25"]),
]

start_row = 5
for i, (per, rec, acad, alunos, tp_freq, tp_rev) in enumerate(HIST_DATA):
    r = start_row + i
    ws2.cell(row=r, column=1, value=per).font = bold
    if rec is not None:
        ws2.cell(row=r, column=2, value=rec).number_format = "#,##0.0"
    ws2.cell(row=r, column=3, value=acad).number_format = "#,##0"
    ws2.cell(row=r, column=4, value=alunos).number_format = "#,##0"
    if i > 0:
        ws2.cell(row=r, column=5, value=f"=(C{r}+C{r-1})/2").number_format = "#,##0.0"
        ws2.cell(row=r, column=6, value=f"=(D{r}+D{r-1})/2").number_format = "#,##0"
        if rec is not None:
            ws2.cell(row=r, column=7, value=f"=B{r}*4/E{r}").number_format = "0.00"
            ws2.cell(row=r, column=9, value=f"=B{r}*1000/3/F{r}").number_format = "0.0"
            ws2.cell(row=r, column=11, value=f"=F{r}/E{r}/1000").number_format = "0.00"
        if tp_freq is not None:
            ws2.cell(row=r, column=17, value=tp_freq / 100).number_format = "0.0%"
            ws2.cell(row=r, column=18, value=tp_rev / 100).number_format = "0.0%"
            ws2.cell(row=r, column=13, value=f"=F{r}/(1-Q{r})/E{r}/1000").number_format = "0.00"
            ws2.cell(row=r, column=15, value=f"=B{r}*(1-R{r})*1000/3/F{r}").number_format = "0.0"
    if i >= 4:
        for col_data, col_yoy in [(7, 8), (9, 10), (11, 12), (13, 14), (15, 16)]:
            cell_curr = f"{get_column_letter(col_data)}{r}"
            cell_prior = f"{get_column_letter(col_data)}{r-4}"
            ws2.cell(row=r, column=col_yoy, value=f'=IFERROR({cell_curr}/{cell_prior}-1,"")').number_format = "+0.0%;-0.0%;0.0%"

# 1T26 row — receita é OUTPUT
r_1t26 = start_row + len(HIST_DATA)
ws2.cell(row=r_1t26, column=1, value="1T26 (est)").font = Font(bold=True, color="C00000")

# Inputs do Premissas
# B4 = # academias EoP, B5 = # alunos EoP, B6 = ticket ex-TP, B10 = TP freq%, B11 = TP rev%
ws2.cell(row=r_1t26, column=3, value="=Premissas!B4").number_format = "#,##0"  # # acad EoP
ws2.cell(row=r_1t26, column=4, value="=Premissas!B5").number_format = "#,##0"  # alunos EoP
ws2.cell(row=r_1t26, column=5, value=f"=(C{r_1t26}+C{r_1t26-1})/2").number_format = "#,##0.0"  # acad média
ws2.cell(row=r_1t26, column=6, value=f"=(D{r_1t26}+D{r_1t26-1})/2").number_format = "#,##0"  # alunos média
ws2.cell(row=r_1t26, column=15, value="=Premissas!B6").number_format = "0.0"  # ticket ex-TP
ws2.cell(row=r_1t26, column=17, value="=Premissas!B10").number_format = "0.0%"  # TP freq%
ws2.cell(row=r_1t26, column=18, value="=Premissas!B11").number_format = "0.0%"  # TP rev%
# Receita = OUTPUT calculado
# Receita_total = (ticket_ex_TP × alunos_média × 3) / (1 - tp_rev) / 1000
ws2.cell(row=r_1t26, column=2, value=f"=O{r_1t26}*F{r_1t26}*3/1000/(1-R{r_1t26})").number_format = "#,##0.0"
# R/Loja
ws2.cell(row=r_1t26, column=7, value=f"=B{r_1t26}*4/E{r_1t26}").number_format = "0.00"
# Ticket mensal (consolidado)
ws2.cell(row=r_1t26, column=9, value=f"=B{r_1t26}*1000/3/F{r_1t26}").number_format = "0.0"
# Alunos/loja
ws2.cell(row=r_1t26, column=11, value=f"=F{r_1t26}/E{r_1t26}/1000").number_format = "0.00"
# Alunos/loja c/TP
ws2.cell(row=r_1t26, column=13, value=f"=F{r_1t26}/(1-Q{r_1t26})/E{r_1t26}/1000").number_format = "0.00"
# YoY 1T26 vs 1T25
for col_data, col_yoy in [(7, 8), (9, 10), (11, 12), (13, 14), (15, 16)]:
    cell_curr = f"{get_column_letter(col_data)}{r_1t26}"
    cell_prior = f"{get_column_letter(col_data)}{r_1t26-4}"
    ws2.cell(row=r_1t26, column=col_yoy, value=f'=IFERROR({cell_curr}/{cell_prior}-1,"")').number_format = "+0.0%;-0.0%;0.0%"

for c in range(1, 19):
    ws2.cell(row=r_1t26, column=c).fill = estimate_fill
    ws2.cell(row=r_1t26, column=c).border = border

widths = [12, 14, 11, 13, 11, 12, 14, 9, 13, 9, 12, 9, 17, 9, 14, 9, 10, 10]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[4].height = 32
ws2.freeze_panes = "B5"

# Save
out = r"C:\Users\diogo.bodas\Desktop\Equity-wiki\equity-wiki\analises\smartfit_kpis_1T26_estimate.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f"Saved: {out}")
